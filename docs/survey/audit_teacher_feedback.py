"""Rebuild a reviewable, question-scoped evidence ledger from archived responses.

This verifies the supplied XLSX against archived transcriptions and recomputes
statistics. It does not verify a later live sheet or independently validate coding.
"""
from __future__ import annotations

import hashlib
import json
import re
import argparse
import statistics
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "RESULTS_2026-07-24_교사시연.md"
OUT = ROOT / "teacher_feedback_audit_2026-07-24.json"
REPORT = ROOT / "자유응답_부호화확정_2026-09-05.md"

CODES = {
    "terms": "용어·기호·단위·영어 표기의 뜻",
    "concepts": "사전 개념·통계 지식·전공 배경 관련 어려움",
    "meaning": "자료·과정·그래프 요소의 의미 설명",
    "independence": "학습자가 근거를 가지고 설명할 기회와 지원",
    "next_action": "단계에서 할 일의 가시성",
    "self_check": "자가점검 문항의 발견과 이용",
    "layout": "화면 구성 단순화",
    "language": "설명 문구의 자연스러움",
    "progression": "쉬운 자료 해석부터의 난이도 단계화",
    "classroom": "교육과정·수업 시간·학생 수준과의 정합",
    "teacher_support": "진행 교사의 사전 설명 준비",
    "concurrency": "다인원 동시 사용의 응답 지연",
    "access": "코딩·자료 다운로드·재구성의 실행 부담",
    "positive": "긍정 평가 또는 활용 의향",
    "choice": "자료 선택과 학습자 주도 수행의 여지",
}

# Every nonempty archived answer is explicitly considered. A code is a review
# judgment, never a count inferred from keyword matching or an AI confidence score.
ANNOTATIONS = {
    "Q6-1": ("학생·수업 예상", ["classroom"]),
    "Q6-2": ("학생·수업 예상", ["classroom"]),
    "Q6-3": ("학생·수업 예상", ["classroom"]),
    "Q6-4": ("학생 예상", ["meaning"]),
    "Q6-5": ("학생 예상", ["concepts", "access"]),
    "Q6-6": ("관점 미명시", ["terms", "meaning"]),
    "Q6-7": ("학생·수업 예상", ["access", "classroom"]),
    "Q6-8": ("학생 예상", ["concepts", "classroom"]),
    "Q6-9": ("교사의 자료 준비", ["access", "classroom"]),
    "Q6-10": ("학생·수업 예상", ["terms", "classroom"]),
    "Q6-11": ("본인 사용 경험과 학생 지원 제안", ["positive", "next_action", "meaning", "terms"]),
    "Q6-12": ("학생·수업 예상", ["classroom", "positive"]),
    "Q20-1": ("본인 경험", ["concepts"]),
    "Q20-3": ("관점 미명시", ["concepts"]),
    "Q20-4": ("경험한 화면에 대한 설명 요구", ["meaning"]),
    "Q20-5": ("학생 예상", ["concepts"]),
    "Q20-9": ("학생·수업 예상", ["concepts", "classroom"]),
    "Q20-10": ("학생 예상", ["terms", "concepts"]),
    "Q20-11": ("학생 예상·현재 단서에 대한 조건부 우려", ["concepts", "classroom", "independence"]),
    "Q22-1": ("학생·수업 예상", ["classroom"]),
    "Q22-2": ("학생 예상", ["terms"]),
    "Q22-3": ("교사의 수업 운영", ["classroom"]),
    "Q22-4": ("학생 활동에 대한 긍정", ["positive"]),
    "Q22-5": ("학생 예상", ["concepts"]),
    "Q22-6": ("학생·수업 예상", ["terms", "concepts"]),
    "Q22-8": ("학생·수업 예상", ["classroom", "progression"]),
    "Q22-9": ("학생·수업 예상", ["classroom", "independence", "choice"]),
    "Q22-10": ("학생·수업 예상", ["terms", "concepts", "classroom"]),
    "Q22-11": ("수업 적용 우려", ["classroom"]),
    "Q22-12": ("교사의 설명 준비와 긍정 평가", ["concepts", "teacher_support", "positive"]),
    "Q23-1": ("본인 평가·보완 제안", ["positive", "layout"]),
    "Q23-5": ("학생 활동에 대한 긍정", ["positive"]),
    "Q23-9": ("본인 평가·보완 제안", ["language", "positive"]),
    "Q23-10": ("본인 평가·수업 운영 요구", ["positive", "concurrency"]),
    "Q23-12": ("학습자 지원의 보완 제안", ["self_check"]),
}

NOTES = {
    "Q6-4": "자료·과정 의미의 이해 어려움이다. 사전 지식 부족을 명시한 응답으로 단정하지 않는다.",
    "Q6-9": "다운로드 후 재구성 부담이 명시돼 있다. 기존 서비스 맥락의 자료 접근 부담을 0으로 보고할 수 없다.",
    "Q6-11": "기존 서비스 문항에 EASWA와 맞물리는 구체 화면 경험이 기입된 예외다. 삭제하지 않고 보충 근거로 보존하되 EASWA 문항의 빈도에 합산하지 않는다.",
    "Q20-1": "본인이 비전공자여서 어렵다는 반응이며 용어 풀이를 직접 요구한 것은 아니다.",
    "Q20-3": "수학·통계 지식의 필요성이다. 특정 용어 풀이 요청이나 본인 경험으로 자동 변환하지 않는다.",
    "Q20-11": "키워드가 현재 없다는 말이 아니다. 제시된 단서 없이 독립적으로 설명할 수 있을지의 우려다.",
    "Q22-9": "안내가 과도하다는 직접 진술이 아니다. 교육과정·수준·자료 선택과 교사 주도 수행에 대한 우려다.",
    "Q23-12": "자가점검 UI가 눈에 띄고 진단 기능이 분명하기를 요구했다. 첫 답 고정이나 특정 채점 방식을 요구하지는 않았다.",
}


def read_answers(text: str) -> dict[str, str]:
    answers = {}
    for question in ("Q6", "Q20", "Q22", "Q23"):
        match = re.search(rf"^### {question} [^\n]*\n(.*?)(?=^### |^> ★|^---|\Z)", text, re.M | re.S)
        assert match, question
        for row in re.finditer(r"^(\d+)\. (.*?)(?=^\d+\. |\Z)", match[1], re.M | re.S):
            answers[f"{question}-{row[1]}"] = row[2].strip()
    assert set(answers) == set(ANNOTATIONS), (set(answers) - set(ANNOTATIONS), set(ANNOTATIONS) - set(answers))
    assert len(answers) == 35
    return answers


def verify_workbook(path: Path, answers: dict[str, str]) -> dict:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book.worksheets[0]
    matrix = list(sheet.iter_rows(values_only=True))
    headers = matrix[0]
    rows = [list(row) + [None] * (len(headers)-len(row)) for row in matrix[1:] if any(v is not None for v in row)]
    assert len(rows) == 13, "Review cohort membership if the source has changed"
    def col(q):
        return next(i for i,h in enumerate(headers) if str(h).startswith(str(q)+'. '))
    cells, nonanswers = {}, []
    for q in ('6','20','22','23'):
        c = col(q)
        for i,row in enumerate(rows[:12],1):
            value = '' if row[c] is None else str(row[c]).strip()
            key = f'Q{q}-{i}'
            address = f'{get_column_letter(c+1)}{i+1}'
            if value and value != '.':
                assert key in answers, key
                assert re.sub(r'\s+',' ',value) == re.sub(r'\s+',' ',answers[key]), key
                cells[key] = address
            else:
                nonanswers.append(dict(locator=key,cell=address,kind='punctuation_only' if value else 'blank',value=value))
    assert len(cells) == 35
    assert sum(r['kind']=='punctuation_only' for r in nonanswers)==4
    q21_options = [
        '자료 출처와 분석 조건을 더 명확히 제시하는 것',
        '분석 과정과 품질 점검 정보를 더 자세히 제공하는 것',
        '그래프와 분석 결과를 해석할 수 있는 도움말을 제공하는 것',
        'STEP별 질문과 생각해보기를 보완하는 것',
        '기준값 비교, 차이 원인 설명, 결과 기록 활동을 강화하는 것',
        '수업 적용을 위한 활동지와 교사용 안내 자료를 제공하는 것',
        '추가 탐구 주제를 제공하는 것',
        '동시에 자료 출력이 가능하게..? 다인원 수용 가능하게',
    ]
    scenarios = {}
    for n in (12,13):
        selected=rows[:n]
        stats={}
        scores=[]
        for q in range(8,19):
            c=col(q); raw=[int(r[c]) for r in selected]
            assert all(1 <= x <= 5 for x in raw)
            oriented=[6-x if q in (12,15) else x for x in raw]
            stats[str(q)]=dict(range=f'{get_column_letter(c+1)}2:{get_column_letter(c+1)}{n+1}',
                              raw_distribution=[raw.count(v) for v in range(1,6)],
                              oriented_mean=statistics.mean(oriented),sample_sd=statistics.stdev(oriented),
                              favorable=sum(v>=4 for v in oriented),neutral=oriented.count(3),unfavorable=sum(v<=2 for v in oriented))
        for row in selected:
            scores.append(statistics.mean(6-int(row[col(q)]) if q in (12,15) else int(row[col(q)]) for q in range(8,19)))
        q21={option:[i+1 for i,r in enumerate(selected) if option in str(r[col('21')] or '')] for option in q21_options}
        # Check whole option strings: the fifth option itself contains commas.
        for r in selected:
            value=str(r[col('21')] or '')
            for option in sorted(q21_options,key=len,reverse=True):
                value=value.replace(option,'')
            assert not value.strip(' ,'), value
        complete=[i+1 for i,r in enumerate(selected) if str(r[col('7')]).startswith('Step 0~6')]
        time_10_30=[i+1 for i,r in enumerate(selected) if r[col('7-1')] in ('10~20분','20~30분')]
        scenarios[str(n)]=dict(n=n,likert=stats,q21=q21,complete=complete,time_10_30=time_10_30,
                              complete_time_10_30=sorted(set(complete)&set(time_10_30)),
                              composite_mean=statistics.mean(scores),composite_sample_sd=statistics.stdev(scores))
    assert scenarios['12']['complete_time_10_30']==[2,7,8,9,10,12]
    return dict(path=str(path.resolve()),sha256=hashlib.sha256(path.read_bytes()).hexdigest(),sheet=sheet.title,
                rows=len(rows),headers={get_column_letter(i+1):h for i,h in enumerate(headers)},
                cells=cells,nonanswers=nonanswers,scenarios=scenarios,
                eligibility='12 archived responses verified; row 14 at 21:45 requires participant-membership confirmation',
                late_response=dict(row=14,timestamp=str(rows[12][0]),Q6=rows[12][col('6')],Q22=rows[12][col('22')]))


def main() -> None:
    parser=argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--xlsx',type=Path,required=True,help='Original Google Forms response export; read only')
    args=parser.parse_args()
    text = SOURCE.read_text(encoding="utf-8-sig")
    answers = read_answers(text)
    workbook=verify_workbook(args.xlsx,answers)
    records, counts = [], defaultdict(lambda: defaultdict(set))
    for locator, answer in answers.items():
        question, respondent = locator.split("-")
        domain = "EASWA" if question != "Q6" else ("EASWA 보충" if locator == "Q6-11" else "기존 서비스")
        perspective, codes = ANNOTATIONS[locator]
        for code in codes:
            assert code in CODES
            counts[domain][code].add(int(respondent))
        records.append(dict(locator=locator, question=question, respondent=int(respondent), domain=domain,
                            source_cell=workbook['cells'][locator],source_sheet=workbook['sheet'],
                            perspective=perspective, codes=codes, note=NOTES.get(locator, ""),
                            text=answer, text_sha256=hashlib.sha256(answer.encode()).hexdigest()))

    # Arithmetic checks use the archived item frequency table, never joint data
    # that cannot be reconstructed from marginal counts.
    quantitative = {}
    for row in re.finditer(r"^\| [▲▼ ]*(\d+)\. .*?\| ([\d.]+) \| \*?\*?([\d.]+)\*?\*? \| ([\d.]+) \| ([\d.]+) \| (\[[^\]]+\]) \|$", text, re.M):
        q = int(row[1]); dist = json.loads(row[6]); n = sum(dist)
        assert n == 12
        raw_mean = sum((i+1)*v for i,v in enumerate(dist))/n
        oriented_mean = 6-raw_mean if q in (12,15) else raw_mean
        assert abs(raw_mean-float(row[2])) <= .00501, q
        assert abs(oriented_mean-float(row[3])) <= .00501, q
        favorable = sum(dist[:2] if q in (12,15) else dist[3:])
        unfavorable = sum(dist[3:] if q in (12,15) else dist[:2])
        quantitative[str(q)] = dict(n=n, raw_distribution=dist, raw_mean=raw_mean,
                                   oriented_mean=oriented_mean, favorable=favorable,
                                   neutral=dist[2], unfavorable=unfavorable)
    assert len(quantitative) == 11
    assert quantitative["15"]["favorable"] == 8 and quantitative["15"]["unfavorable"] == 3
    assert counts["EASWA"]["terms"] == {2,6,10}
    assert counts["EASWA 보충"]["terms"] == {11}

    for q,stats in quantitative.items():
        assert stats['raw_distribution']==workbook['scenarios']['12']['likert'][q]['raw_distribution']
    result = dict(status="original_xlsx_verified_12_response_baseline_late_response_eligibility_pending",
                  original_workbook=workbook,
                  source=SOURCE.name, source_sha256=hashlib.sha256(SOURCE.read_bytes()).hexdigest(),
                  scope="35 nonempty answers: Q6=12, Q20=7, Q22=11, Q23=5; Q20-1 excluded as performance artifact",
                  codes=CODES, records=records, quantitative=quantitative,
                  counts={d:{c:sorted(ids) for c,ids in cs.items()} for d,cs in counts.items()})
    OUT.write_text(json.dumps(result,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")

    lines = ["# 현직교사 검토 재대조 — 문항 맥락을 보존한 부호화 검토본", "",
             "> 2026-09-05. 종전의 ‘사전 지식·용어 11명 확정’ 해석을 이 검토본으로 교체한다.",
             "> 원본 XLSX의 기존 12명 응답과 서술형 전사본 35개를 대조했다. 빈도·역채점은 일치하며, 21:45 추가 응답 1개의 시연 참가 여부 확인이 남아 있다.",
             "> 이 파일은 독립 부호화자 간 신뢰도 확인이나 연구자의 최종 부호화 승인을 뜻하지 않는다.", "",
             f"원자료: [{SOURCE.name}]({SOURCE.name}) §8. 응답 번호는 저장 전사본의 번호를 유지한다.",
             f"기계 확인·응답별 전문: [{OUT.name}]({OUT.name}). 재생성: `python -X utf8 docs/survey/audit_teacher_feedback.py --xlsx <원본.xlsx>`.", "",
             "## 1. 집계 전에 고정한 구분", "",
             "- Q5·Q6은 기존 서비스의 경험·예상 장벽이다. EASWA 사용 후 요구의 빈도에 합치지 않는다.",
             "- Q20·Q22·Q23은 EASWA 관련 의견이지만 본인 경험·학생 예상·수업 운영 요구를 구분한다.",
             "- Q6-11은 EASWA와 맞물리는 화면·용어를 구체적으로 언급한 예외다. 보충 근거로 사용하고 별도 표기한다.",
             "- 빈도는 같은 조사 대상·범주 안의 고유 응답자 수다. 복수선택 인원과 자유응답 인원을 더하지 않는다.",
             "- 어려움·조건부 우려·직접 수정 요구를 같은 의미로 바꾸지 않는다. 코드 없는 무응답은 어려움 없음으로 판정하지 않는다.", "",
             "## 2. EASWA 문항에서 직접 추적되는 범주", "",
             "| 범주 | 응답자 번호 | 수 | Q6-11 보충 |", "|---|---|---:|---|"]
    for code,label in CODES.items():
        ids=sorted(counts["EASWA"].get(code,set()))
        extra="별도 1건" if code in counts["EASWA 보충"] else "—"
        lines.append(f"| {label} | {'·'.join(map(str,ids)) or '—'} | {len(ids)} | {extra} |")
    lines += ["", "범주 폭이 다르므로 이 표의 순위를 그대로 수정 우선순위로 사용하지 않는다. ‘사전 개념·배경’은 용어 풀이만으로 해결되는 요구가 아니다.", "",
              "## 3. 철회하거나 좁혀야 할 기존 해석", "",
              "| 기존 해석 | 원문·수치 대조 | 개선 도출에 미치는 영향 |", "|---|---|---|",
              "| 사전 지식·용어 11명이 EASWA 용어 보완의 최다 근거 | 기존 서비스 Q6과 EASWA 문항, 용어와 전공 배경을 합쳤다. EASWA 문항의 명시적 용어 언급은 3명이며 Q6-11이 별도 보충 근거다. | 용어·개념·그래프 읽기·난이도 지원을 나누어 대응한다. |",
              "| Q15 최하위이므로 해석 지원이 실패 | 역문항 원점수 [1,7,1,2,1]은 어려움 진술에 반대 8명·중립 1명·동의 3명이다. | 긍정 응답이 다수여도 Q21 각 7명과 구체 서술은 보완 근거로 남는다. |",
              "| 학생의 해석·기록 어려움 9명이 실제 사용 결함 | 9명은 학생 예상이며 본인 경험은 2명이다. | 교사의 수업 설계 우려와 직접 사용 문제를 구분하고 2차는 본인 경험을 별도 측정한다. |",
              "| Q22-9는 안내 과잉 | 교사 설명과 버튼 따라 누르기에 그칠 우려이며, 안내를 줄여 달라는 진술이 없다. | 필요한 안내를 유지하면서 학습자의 선택·예측·근거 설명 기회를 확보한다. |",
              "| 활동지는 앱 밖이므로 제외 | Q21에서 6명이 선택했고 서론 1.4·원리 5가 교사 지원 자료를 요구한다. | 최소 활동 안내·교사 준비 자료를 개선 목록에 남긴다. 구현하지 않은 상태는 그대로 기록한다. |",
              "| 자료 접근 부담 0 | Q6-9에 자료 다운로드와 재구성 부담이 명시되어 있다. | 기존 서비스 분석에서 삭제하지 않는다. EASWA 내부 다운로드 오류로 바꾸지도 않는다. |",
              "| 도움말 요구 7명과 학생 주도성 우려가 정면 충돌 | 이해를 위한 안내와 교사 주도 대행은 다른 내용이다. | 안내의 양 대신 정보 종류·배치·학습자 수행을 점검한다. |",
              "| 동아리 5명이므로 정규 차시보다 선호 | 1차시 1명·2차시 4명·3차시 이상 2명, 동아리 5명이다. | 동아리는 단일 최빈 선택지일 뿐 정규 차시 응답 전체 7명보다 많지 않다. |",
              "| 완주 9명 중 8명이 10~30분 | 원본 J2:K13 교차 확인: 완주 9명 중 10~30분은 6명(응답자 2·7·8·9·10·12)이다. 8명은 전체 12명의 시간 집계다. | ‘완주자 가운데 6명’ 또는 두 문항의 별도 집계로 고친다. |",
              "| Q21에 보완 요소 언급이 없으면 낮은 배경지식 탓 | 언급 부재는 원인의 증거가 아니다. | 보완 노출·사용 경험·구체 막힘을 묻고 원인이 불명확하면 미확인으로 남긴다. |", "",
              "## 4. 실제 개선과 2차 측정을 연결하는 작업표", "",
              "| 개선 항목 | 직접 근거 | 서론·원리 연결 | 구현에서 확인할 것 | 2차에서 확인할 것 |", "|---|---|---|---|---|",
              "| 용어·기호·단위와 선수 개념 설명 | Q22-2·6·10, Q20-9·10, 보충 Q6-11 | 1.2·1.4 / 원리 3·5 | BTJD 외 BJD·ROI·기호·단위가 실제 표시되는 지점의 풀이, 차등측광 등 짧은 사전 개념 | 화면 설명으로 이해했는지, 사람의 도움이 필요했는지, 이해 안 된 용어 |",
              "| 그래프 읽기 | Q20-4, 보충 Q6-11; Q21 해석 도움말 7명 | 1.4 / 원리 3·4 | 축·점·빈 구간·품질 그래프의 의미, 화면에 실제 제시되는 값 | 의미를 설명할 수 있었는지와 막힌 요소; 특정 원인 정답을 미리 주지 않음 |",
              "| 각 단계의 할 일·자가점검 가시성 | 보충 Q6-11, Q23-12, Q23-1; Q21 STEP 보완 4명 | 1.4 / 원리 1·5 | 다음 행동이 보이는지, 질문 위치를 찾는지, 중복 설명·중복 기록 | Q19 본인 막힘과 Q20 구체 위치·이유; 새 보충 문항은 별도 보고 |",
              "| 근거를 사용한 해석·기록 | Q21 기준값 비교·원인·기록 7명, Q20-11, Q22-9 | 1.4 / 원리 4 | 실제 사용한 조건을 근거로 선택·예측·설명하는 기회, 원인 정답 대행 여부 | Q20-1 원인 서술과 근거 연결; 정답률만으로 독립적 해석을 주장하지 않음 |",
              "| 최소 활동 안내·교사 준비 자료 | Q21 활동지·교사 안내 6명, Q22-12; Q22-8 난이도 단계화 | 1.4·표 2-1 / 원리 5 | 학습 목표·선수 개념·필수/확장 활동·진행 팁; 실제 제공 여부 | 실제 자료를 본 경우만 내용·분량의 적절성; 보지 않았다면 미노출 |",
              "| 문체와 동시 사용 | Q23-9, Q23-10 | 1.4 / 원리 2·5 | 부자연스러운 문장 교체, 실제 동시 수행 시 지연 | 문체 관련 구체 의견과 실제 지연 기록; 일반 인프라 과제로만 밀지 않음 |", "",
              "위 작업표는 요구 도출이다. 화면 기능의 구현 완료나 배포 완료를 인증하지 않는다. 첫 답 고정 같은 기록 수집 결함은 앱 점검 근거로 별도 관리하며 교사가 직접 요구한 수정으로 서술하지 않는다.", "",
              "## 5. 응답별 재대조 — 35개 전수", "",
              "| 원문 위치 / XLSX 셀 | 대상 | 관점 | 부호 | 원문 | 해석 주의 |", "|---|---|---|---|---|---|"]
    for r in records:
        clean=lambda s:s.replace("|","\\|").replace("\n","<br>")
        lines.append("| "+" | ".join(clean(s) for s in [r["locator"]+' / '+r['source_cell'],r["domain"],r["perspective"]," / ".join(CODES[c] for c in r["codes"]),r["text"],r["note"]])+" |")
    lines += ["", "## 6. 검증 범위", "",
              "- 저장 전사본 35개의 위치·전문·고유 응답자 집계를 확인했다.",
              "- 척도 11문항의 빈도 합(각 12)과 원평균·역채점평균을 재계산했다.",
              f"- 원본: `{workbook['path']}` / 시트 `{workbook['sheet']}`. 파일 SHA-256: `{workbook['sha256']}`.",
              "- 기존 12명은 XLSX 2~13행이다. 서술형의 공백만 정규화하여 35개 전문과 응답자 번호의 일치를 확인했다. 마침표만 적은 4개는 내용 없는 응답으로 별도 기록했다.",
              "- 14행은 같은 날 21:45 제출된 동의 응답이다. 시연 참가자의 지연 제출인지 확인 전이므로 삭제하거나 주 분석에 자동 편입하지 않았다. JSON에는 N=12와 N=13 통계를 모두 계산했다.",
              "- 부호화는 원문을 읽고 작성한 검토안이다. 연구자의 최종 범주 승인·독립 재부호화를 대신하지 않는다.", ""]
    REPORT.write_text("\n".join(lines),encoding="utf-8")
    print(json.dumps(dict(answers=len(records),likert_items=len(quantitative),easwa_terms=sorted(counts["EASWA"]["terms"]),supplemental_terms=sorted(counts["EASWA 보충"]["terms"]),report=str(REPORT),status=result["status"]),ensure_ascii=False))


if __name__ == "__main__":
    main()
